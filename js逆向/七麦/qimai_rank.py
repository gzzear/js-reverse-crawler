"""
七麦数据 - App Store 榜单爬取脚本
目标页面: https://www.qimai.cn/rank/index/brand/free/device/iphone/country/cn/genre/36

【逆向分析说明】
1. 网站为 Vue SPA，前端路由在客户端处理，未登录时路由守卫重定向到 /404
2. 但在重定向前，API 请求已经发出并返回数据，可直接调用
3. 真实 API 端点: https://api.qimai.cn/rank/indexPlus/brand_id/{brand_id}
   - brand_id=0: 付费榜
   - brand_id=1: 免费榜
   - brand_id=2: 畅销榜
4. `analysis` 参数: 由前端 JS 动态生成的签名 token，已完全逆向，本脚本可自动生成
5. 分页: 通过 ?page=N 参数翻页，每页 20 条
6. 未登录限制: 最多返回 210 条 (11页)；VIP 账号可获取完整 500 条

【analysis token 生成算法（已逆向）】
XOR_KEY = "6efghxyz517cda9"（15字符循环）
明文结构 = cv(sorted_param_values) + "@#" + url_path + "@#" + timeDelta + "@#3"
  - cv(x)    = base64(x)
  - url_path = 去掉 baseURL 后的路径，如 /rank/indexPlus/brand_id/1
  - timeDelta = int(time.time()*1000) + syncd_ms - 1661224081041
  - syncd_ms = 浏览器 cookie 中 syncd 的整数值（负数），表示服务器与客户端时钟偏差
analysis = base64(XOR(明文, XOR_KEY))

【genre_id 与榜单类型对应】
genre_id 参数加入 URL path 后对未登录用户无效（返回综合榜），
需登录后才能按分类过滤（音乐=36等为七麦内部编号）
"""

import base64
import requests
import openpyxl
import time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────
# 配置区
# ────────────────────────────────────────────────

# analysis token 自动生成所需参数
# syncd: 浏览器 cookie 中 syncd 的值（整数，通常为负数，如 -1452009）
# 表示服务器时钟与客户端时钟的偏差（毫秒），从 https://www.qimai.cn 的 cookie 中获取
SYNCD_MS = -218427

# 登录后的 Cookie（未登录只能获取前210条，登录后可获取500条）
COOKIES = {
    "qm_check": "A1sdRUIQChtxen8pI0dAOQkKWVIeEHh+c3QgRioNDBgWFWVXXl1VRl0XXEcpCAkWUBd/ARUQYVYWFgILER8TUFMSZlxCR1EKCE5KVFsZXVJRWxsKFghJVktYVElWBRVP",
    "PHPSESSID": "ns164esn7ku04n9d4i7ajsjgj6"
}

# ────────────────────────────────────────────────
# analysis token 自动生成
# ────────────────────────────────────────────────

_XOR_KEY = "6efghxyz517cda9"
_BASE_TS = 1661224081041


def _xor(text: str, key: str) -> bytes:
    kb = key.encode()
    return bytes(b ^ kb[i % len(kb)] for i, b in enumerate(text.encode()))


def generate_analysis(url_path: str, params: dict) -> str:
    """
    自动生成七麦 API 的 analysis 签名 token。

    算法（已逆向自 webpack 模块 65165 的 axios 请求拦截器）：
      plaintext = base64(sorted_param_values) + "@#" + url_path + "@#" + timeDelta + "@#3"
      analysis  = base64(XOR(plaintext, "6efghxyz517cda9"))
    """
    # 排序参数值（排除 analysis 本身）
    values = sorted(str(v) for k, v in params.items() if k != "analysis")
    cv = base64.b64encode("".join(values).encode()).decode()

    time_delta = int(time.time() * 1000) + SYNCD_MS - _BASE_TS

    plaintext = f"{cv}@#{url_path}@#{time_delta}@#3"
    return base64.b64encode(_xor(plaintext, _XOR_KEY)).decode()


# 榜单配置
BRAND_CONFIG = {
    "free": {"brand": "free", "name": "免费榜"},
    "paid": {"brand": "paid", "name": "付费榜"},
    "grossing": {"brand": "grossing", "name": "畅销榜"},
}

TARGET_BRAND = "paid"  # 目标榜单类型
TARGET_GENRE = 5000  # 七麦内部 genre_id（仅登录后有效）
TARGET_DEVICE = "iphone"
TARGET_COUNTRY = "us"
MAX_PAGES = 30  # 最大抓取页数（防止死循环）
PAGE_DELAY = 0.3  # 每页请求间隔（秒），避免被限速

BASE_API = "https://api.qimai.cn"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
    "Referer": "https://www.qimai.cn/",
    "Origin": "https://www.qimai.cn",
    "Accept": "application/json, text/plain, */*",
}


def fetch_rank_page(brand: str, page: int, genre: int,
                    device: str = "iphone", country: str = "cn") -> list:
    """
    获取单页榜单数据。
    API 端点: GET /rank/index?brand=free&device=iphone&country=cn&genre=5000&page=1
    """
    url_path = "/rank/index"
    params = {"brand": brand, "device": device, "country": country, "genre": genre, "page": page, "date": "2026-03-23"}
    params["analysis"] = generate_analysis(url_path, params)

    resp = requests.get(BASE_API + url_path, params=params, headers=HEADERS, cookies=COOKIES, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    print(data)
    if data.get("code") != 10000:
        raise RuntimeError(f"API 返回错误: {data}")

    return data.get("rankInfo", data.get("list", []))


def fetch_all_pages(brand: str, genre: int) -> list:
    """分页获取所有榜单数据，直到返回空列表为止。"""
    all_items = []
    for page in range(1, MAX_PAGES + 1):
        items = fetch_rank_page(brand, page, genre, TARGET_DEVICE, TARGET_COUNTRY)
        if not items:
            break
        all_items.extend(items)
        print(f"  第 {page} 页: {len(items)} 条，累计 {len(all_items)} 条")
        time.sleep(PAGE_DELAY)
    return all_items


def parse_item(item: dict) -> dict:
    """将 API 返回的单条数据解析为扁平化结构。"""
    info = item.get("appInfo", {})
    size_bytes = info.get("file_size", 0)
    try:
        size_mb = round(int(size_bytes) / 1024 / 1024, 1)
    except (ValueError, TypeError):
        size_mb = ""

    # /rank/index 的 change 在 rank_a.change 里；旧接口在 item.change
    rank_a = item.get("rank_a", {})
    raw_change = rank_a.get("change", item.get("change", 0))
    try:
        raw_change = int(raw_change)
        change_str = f"+{raw_change}" if raw_change > 0 else str(raw_change)
    except (ValueError, TypeError):
        change_str = str(raw_change)

    # 分类：新接口在 rank_a.genre，旧接口在 appGenre/genre
    genre_str = rank_a.get("genre") or item.get("appGenre") or item.get("genre", "")

    price = info.get("price", "0.00")
    price_str = "免费" if price == "0.00" else f"¥{price}"

    return {
        "排名": item.get("index", ""),
        "应用名称": info.get("appName", ""),
        "副标题": info.get("subtitle", ""),
        "分类": genre_str,
        "开发者": info.get("publisher", ""),
        "价格": price_str,
        "排名变化": change_str,
        "AppID": info.get("appId", ""),
        "是否广告": "是" if item.get("is_ad") else "否",
        "文件大小MB": size_mb,
        "图标链接": info.get("icon", ""),
    }


def save_to_excel(rows: list, output_path: str, sheet_name: str = "榜单"):
    """将数据保存为格式化的 Excel 文件。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1565C0")
    thin_side = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    alt_fill = PatternFill("solid", fgColor="F5F7FA")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # 列定义: (列名, 宽度, 居中?)
    COLUMNS = [
        ("排名", 6, True),
        ("应用名称", 30, False),
        ("副标题", 30, False),
        ("分类", 10, True),
        ("开发者", 35, False),
        ("价格", 8, True),
        ("排名变化", 10, True),
        ("AppID", 14, True),
        ("是否广告", 8, True),
        ("文件大小MB", 12, True),
        ("图标链接", 60, False),
    ]

    # 写表头
    ws.append([col[0] for col in COLUMNS])
    for i, cell in enumerate(ws[1]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    # 写数据行
    for row_idx, row_data in enumerate(rows, start=2):
        ws.append([row_data.get(col[0], "") for col in COLUMNS])
        is_alt = (row_idx % 2 == 0)
        for col_idx, cell in enumerate(ws[row_idx]):
            cell.border = cell_border
            cell.alignment = center_align if COLUMNS[col_idx][2] else left_align
            if is_alt:
                cell.fill = alt_fill

    # 列宽 & 行高
    for i, (_, width, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22

    # 冻结首行 & 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"\nExcel 已保存: {output_path}")


def main():
    brand = BRAND_CONFIG[TARGET_BRAND]
    print(
        f"开始抓取七麦数据 [{brand['name']}] genre={TARGET_GENRE} device={TARGET_DEVICE} country={TARGET_COUNTRY} ...")

    raw_items = fetch_all_pages(brand["brand"], TARGET_GENRE)
    print(f"\n共获取 {len(raw_items)} 条，开始解析并生成 Excel ...")

    parsed = [parse_item(item) for item in raw_items]

    timestamp = time.strftime("%Y%m%d_%H%M")
    output_path = f"/Users/gaozhe/Desktop/七麦_{brand['name']}_{timestamp}.xlsx"
    save_to_excel(parsed, output_path, sheet_name=f"{brand['name']} Top{len(parsed)}")


if __name__ == "__main__":
    main()
