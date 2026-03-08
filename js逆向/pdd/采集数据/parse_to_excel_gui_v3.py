import glob
import json
import os
import re
import threading
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk, scrolledtext

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# --- 工业级视觉规范 ---
COLOR_NAVY_DARK = "#001529"  # 导航深蓝
COLOR_ANT_BLUE = "#1890ff"  # 品牌蓝
COLOR_BG_LIGHT = "#f0f2f5"  # 页面背景
COLOR_WHITE = "#ffffff"  # 卡片背景

# 极速连接池
http_session = requests.Session()
adapter = requests.adapters.HTTPAdapter(pool_connections=100, pool_maxsize=100)
http_session.mount('https://', adapter)
http_session.mount('http://', adapter)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"}


class ProfessionalLogger:
    def __init__(self, text_widget):
        self.widget = text_widget
        self.widget.tag_config("TIME", foreground="#8c8c8c")
        self.widget.tag_config("INFO", foreground="#1890ff")
        self.widget.tag_config("SUCCESS", foreground="#52c41a", font=("Segoe UI", 9, "bold"))
        self.widget.tag_config("WARNING", foreground="#faad14")
        self.widget.tag_config("ERROR", foreground="#ff4d4f", font=("Segoe UI", 9, "bold"))

    def log(self, message, level="INFO"):
        now = datetime.now().strftime("%H:%M:%S")
        self.widget.config(state=tk.NORMAL)
        self.widget.insert(tk.END, f"[{now}] ", "TIME")
        self.widget.insert(tk.END, f"[{level}] ", level)
        self.widget.insert(tk.END, f"{message}\n")
        self.widget.see(tk.END)
        self.widget.config(state=tk.DISABLED)

    def clear(self):
        self.widget.config(state=tk.NORMAL)
        self.widget.delete(1.0, tk.END)
        self.widget.config(state=tk.DISABLED)


def clean_json_content(content):
    """鲁棒性清洗 JSON 杂质"""
    content = content.strip()
    match = re.search(r'(\{.*\})', content, re.DOTALL)
    return match.group(1) if match else content


class PddParserV2:
    def __init__(self, root):
        self.root = root
        self.root.title("PDD 数据解析专家 v3.0 (网络链接版)")
        self.root.geometry("1000x700")
        self.root.configure(bg=COLOR_BG_LIGHT)

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass

        self.setup_styles()
        self.build_ui()
        self.logger = ProfessionalLogger(self.log_area)

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Card.TFrame", background=COLOR_WHITE, relief="flat")
        style.configure("TLabel", background=COLOR_WHITE, font=("Microsoft YaHei", 10))
        style.configure("Header.TLabel", font=("Microsoft YaHei", 18, "bold"), background=COLOR_NAVY_DARK,
                        foreground="white")
        style.configure("Primary.TButton", font=("Microsoft YaHei", 10, "bold"), padding=8)
        style.map("Primary.TButton", background=[('pressed', '#096dd9'), ('active', COLOR_ANT_BLUE)],
                  foreground=[('!disabled', 'white')])

    def build_ui(self):
        # 1. 顶部标题栏
        nav = tk.Frame(self.root, bg=COLOR_NAVY_DARK, height=65)
        nav.pack(fill=tk.X)
        nav.pack_propagate(False)
        ttk.Label(nav, text=" PDD DATA PARSER V3 ", style="Header.TLabel").pack(side=tk.LEFT, padx=30)
        ttk.Label(nav, text=" 网络链接版 | 不下载图片 ", font=("Microsoft YaHei", 9), background=COLOR_NAVY_DARK,
                  foreground="#595959").pack(side=tk.LEFT, pady=(8, 0))

        # 2. 主卡片容器
        container = ttk.Frame(self.root, padding=30)
        container.pack(fill=tk.BOTH, expand=True)

        cfg_card = ttk.Frame(container, style="Card.TFrame", padding=25)
        cfg_card.pack(fill=tk.X, pady=(0, 20))

        # 输入路径配置
        f1 = ttk.Frame(cfg_card, style="Card.TFrame")
        f1.pack(fill=tk.X)
        ttk.Label(f1, text="数据源路径", font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.in_path = tk.StringVar()
        ttk.Entry(f1, textvariable=self.in_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=20)

        btn_box = ttk.Frame(f1, style="Card.TFrame")
        btn_box.pack(side=tk.LEFT)
        ttk.Button(btn_box, text="单文件", width=8, command=self.select_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_box, text="文件夹", width=8, command=self.select_dir).pack(side=tk.LEFT)

        # 导出路径配置
        f2 = ttk.Frame(cfg_card, style="Card.TFrame")
        f2.pack(fill=tk.X, pady=(20, 0))
        ttk.Label(f2, text="导出 Excel  ", font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.out_path = tk.StringVar()
        ttk.Entry(f2, textvariable=self.out_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=20)
        ttk.Button(f2, text="浏览...", width=10, command=self.select_out).pack(side=tk.LEFT)

        # 3. 控制面板
        ctrl_fm = ttk.Frame(container)
        ctrl_fm.pack(fill=tk.X, pady=(10, 20))
        self.run_btn = ttk.Button(ctrl_fm, text="🚀 开始解析（网络链接版）", style="Primary.TButton",
                                  command=self.start_task)
        self.run_btn.pack(side=tk.LEFT)
        ttk.Button(ctrl_fm, text="🗑 清空日志", command=lambda: self.logger.clear()).pack(side=tk.RIGHT)

        # 4. 进度条
        progress_frame = ttk.Frame(container)
        progress_frame.pack(fill=tk.X, pady=(10, 5))
        
        ttk.Label(progress_frame, text="解析进度:", font=("Microsoft YaHei", 9)).pack(anchor=tk.W)
        self.pbar = ttk.Progressbar(progress_frame, mode='determinate')
        self.pbar.pack(fill=tk.X, pady=(2, 10))
        self.progress_text = tk.StringVar(value="就绪")
        ttk.Label(progress_frame, textvariable=self.progress_text, font=("Microsoft YaHei", 8), 
                 foreground="#8c8c8c").pack(anchor=tk.W)

        # 5. 日志区
        self.log_area = scrolledtext.ScrolledText(container, height=18, font=("Consolas", 10), bg="#ffffff",
                                                  borderwidth=0, highlightthickness=1, highlightbackground="#d9d9d9",
                                                  state=tk.DISABLED)
        self.log_area.pack(fill=tk.BOTH, expand=True, pady=10)

    def select_file(self):
        p = filedialog.askopenfilename(filetypes=[("数据包", "*.txt *.json"), ("所有文件", "*.*")])
        if p: self.in_path.set(p)

    def select_dir(self):
        p = filedialog.askdirectory()
        if p: self.in_path.set(p)

    def select_out(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p: self.out_path.set(p)

    def start_task(self):
        if not self.in_path.get() or not self.out_path.get():
            messagebox.showwarning("提示", "请先配置输入源和保存位置")
            return
        self.run_btn.config(state=tk.DISABLED)
        threading.Thread(target=self.work_thread, daemon=True).start()

    def work_thread(self):
        self.logger.log(">>> 任务启动，开始扫描...", "INFO")
        inp = self.in_path.get()

        # 兼容单文件与文件夹
        if os.path.isfile(inp):
            target_files = [inp]
        else:
            files = glob.glob(os.path.join(inp, "**", "*.*"), recursive=True)
            target_files = [f for f in files if f.lower().endswith(('.json', '.txt'))]

        if not target_files:
            self.logger.log("错误：未发现有效文件包", "ERROR")
            self.run_btn.config(state=tk.NORMAL)
            return

        self.pbar['maximum'] = len(target_files)
        self.pbar['value'] = 0
        data_rows = []

        for idx, f_path in enumerate(target_files):
            # 更新文件解析进度
            self.progress_text.set(f"正在解析: {os.path.basename(f_path)} ({idx + 1}/{len(target_files)})")

            # 智能解析
            res = self.smart_parse(f_path)
            if res:
                data_rows.append(res)
                self.logger.log(f"解析成功: {os.path.basename(f_path)}", "SUCCESS")
            else:
                self.logger.log(f"跳过非详情包: {os.path.basename(f_path)}", "WARNING")

            self.pbar['value'] = idx + 1
            self.root.update_idletasks()

        if data_rows:
            self.logger.log(f"解析完毕。有效数据: {len(data_rows)} / 总文件: {len(target_files)}", "INFO")
            self.progress_text.set(f"解析完成，开始生成Excel...")
            self.generate_excel(data_rows)
        else:
            self.logger.log("解析结果为空，任务提前结束", "ERROR")
            self.progress_text.set("解析失败")
        self.run_btn.config(state=tk.NORMAL)

    def smart_parse(self, path):
        """核心：智能适配多种 PDD 接口结构 - V3网络链接版"""
        content = None
        for enc in ['utf-8-sig', 'utf-8', 'gbk', 'utf-16', 'latin1']:
            try:
                with open(path, 'r', encoding=enc) as f:
                    raw_content = f.read()
                    content = json.loads(clean_json_content(raw_content))
                    break
            except:
                continue
        if not content: return None

        # 智能层级探测（兼容两种格式）
        # 格式1: store.initDataObj.goods
        # 格式2: 直接 goods
        store_data = content.get('store', {})
        init_data = store_data.get('initDataObj', {})
        goods = init_data.get('goods') or content.get('goods')
        if not goods: return None

        def sget(d, *keys):
            for k in keys:
                if k in d: return d[k]
            return None

        # 基础字段
        goods_name = sget(goods, 'goodsName', 'goods_name', 'shareDesc', 'share_desc') or 'NA'
        goods_id = sget(goods, 'goodsID', 'goods_id') or '0'
        cat_id = str(sget(goods, 'catID', 'cat_id') or '')

        # 商品描述：从属性中构建
        props = goods.get('goodsProperty', []) or goods.get('goods_property', [])
        desc_lines = []
        for p in props:
            if isinstance(p, dict) and p.get('key'):
                key = p.get('key')
                values = p.get('values', [])
                if values:
                    value = values[0] if isinstance(values, list) else values
                    desc_lines.append(f"{key}：{value}")
        goods_desc = '\n'.join(desc_lines)

        # SKU处理（兼容两种字段名）
        skus = goods.get('skus', []) or content.get('sku', [])

        # 提取规格维度
        spec_keys = {}  # {spec_key_id: spec_key_name}
        spec_values = {}  # {spec_key_id: [spec_values]}

        for sku in skus:
            if not isinstance(sku, dict):
                continue
            specs = sku.get('specs', [])
            for spec in specs:
                if isinstance(spec, dict):
                    spec_key_id = spec.get('spec_key_id') or spec.get('specKeyId')
                    spec_key = spec.get('spec_key') or spec.get('specKey', '')
                    spec_value = spec.get('spec_value') or spec.get('specValue', '')

                    if spec_key_id:
                        spec_keys[spec_key_id] = spec_key
                        if spec_key_id not in spec_values:
                            spec_values[spec_key_id] = []
                        if spec_value and spec_value not in spec_values[spec_key_id]:
                            spec_values[spec_key_id].append(spec_value)

        # 获取规格一和规格二
        spec_key_ids = sorted(spec_keys.keys())
        spec1_name = spec_keys.get(spec_key_ids[0], '') if len(spec_key_ids) > 0 else ''
        spec1_values = '||'.join(spec_values.get(spec_key_ids[0], [])) if len(spec_key_ids) > 0 else ''
        spec2_name = spec_keys.get(spec_key_ids[1], '') if len(spec_key_ids) > 1 else ''
        spec2_values = '||'.join(spec_values.get(spec_key_ids[1], [])) if len(spec_key_ids) > 1 else ''

        # 价格、库存、规格图片（按SKU顺序）
        prices = []
        stocks = []
        sku_images = []

        for sku in skus:
            if not isinstance(sku, dict):
                continue

            # 价格处理（兼容多种格式）
            group_price = sku.get('groupPrice') or sku.get('group_price')
            sku_price = sku.get('skuPrice') or sku.get('sku_price', 0)

            # 判断价格格式
            if isinstance(group_price, str) and group_price:
                # 已经是字符串格式的元（如 "69.9"）
                price = group_price
            elif isinstance(group_price, (int, float)) and group_price > 0:
                # 数字格式，判断是否需要除以100
                if group_price > 1000:  # 大于1000认为是分
                    price = str(float(group_price) / 100)
                else:
                    price = str(float(group_price))
            elif sku_price:
                # 使用sku_price，通常是分
                price = str(float(sku_price) / 100) if sku_price else '0'
            else:
                price = '0'

            prices.append(price)

            # 库存
            quantity = sku.get('quantity', 0)
            stocks.append(str(quantity))

            # SKU图片URL
            thumb_url = sku.get('thumbUrl') or sku.get('thumb_url', '')
            sku_images.append(thumb_url)

        price_str = '||'.join(prices)
        stock_str = '||'.join(stocks)

        # 主图URL提取
        m_urls = goods.get('viewImageData', [])
        if not m_urls:
            # 尝试从 gallery 提取
            gallery = goods.get('gallery', [])
            if gallery:
                m_urls = [i.get('url') if isinstance(i, dict) else i for i in gallery]
            else:
                # 尝试从 topGallery
                tg = goods.get('topGallery', [])
                m_urls = [i.get('url') if isinstance(i, dict) else i for i in tg]

        # 如果还是没有，尝试从 decoration 提取
        if not m_urls:
            decoration = goods.get('decoration', [])
            for dec in decoration[:6]:
                if isinstance(dec, dict):
                    contents = dec.get('contents', [])
                    if contents and isinstance(contents[0], dict):
                        img_url = contents[0].get('img_url') or contents[0].get('imgUrl')
                        if img_url:
                            m_urls.append(img_url)

        # 最后尝试 thumb_url
        if not m_urls and sget(goods, 'thumbUrl', 'thumb_url', 'hd_thumb_url'):
            m_urls = [sget(goods, 'thumbUrl', 'thumb_url', 'hd_thumb_url')]

        # 重量提取（优先级：goods.weight > skus[].weight > goodsProperty）
        weight = '0'
        goods_weight = goods.get('weight', 0)
        if goods_weight and goods_weight != 0:
            weight = str(goods_weight)
        else:
            # 尝试从SKU中获取（取第一个非0的）
            for sku in skus:
                sku_weight = sku.get('weight', 0)
                if sku_weight and sku_weight != 0:
                    weight = str(sku_weight)
                    break

            # 如果还是0，尝试从goodsProperty中提取
            if weight == '0':
                for prop in props:
                    if '重量' in prop.get('key', ''):
                        values = prop.get('values', [])
                        if values and values[0] not in ['其他', '其他）', '']:
                            weight = str(values[0])
                            break

        # 站点/国家提取（优先级：goods.country > goodsProperty产地 > 默认CN）
        site = 'CN'  # 默认中国

        # 方法1：从 goods.country 获取
        country = goods.get('country', '')
        if country and country.strip():
            site = country.strip()
        else:
            # 方法2：从 goodsProperty 中提取产地或发货地
            for prop in props:
                key = prop.get('key', '')
                if '产地' in key or '发货地' in key:
                    values = prop.get('values', [])
                    if values and values[0]:
                        origin = values[0]
                        # 简单映射：中国大陆->CN, 其他国家可以扩展
                        if '中国' in origin or '大陆' in origin:
                            site = 'CN'
                        elif '泰国' in origin:
                            site = 'TH'
                        elif '美国' in origin:
                            site = 'US'
                        elif '日本' in origin:
                            site = 'JP'
                        elif '韩国' in origin:
                            site = 'KR'
                        else:
                            # 如果是省份（如"河北省"），保持CN
                            if '省' in origin or '市' in origin:
                                site = 'CN'
                            else:
                                site = origin[:2].upper()  # 取前两个字符作为代码
                        break

        # 整理图片URL字符串
        sku_image_url_str = '||'.join([img for img in sku_images if img])
        main_images_url_str = '||'.join([url for url in m_urls if url])
        
        # 构建行数据（V3版：图片列直接放URL，不下载）
        row = [
            str(goods_id),          # 产品ID
            goods_name,             # 标题
            goods_desc,             # 商品描述
            cat_id,                 # 分类ID
            price_str,              # 价格（||分隔）
            spec1_name,             # 规格一
            spec1_values,           # 规格一内容
            spec2_name,             # 规格二
            spec2_values,           # 规格二内容
            sku_image_url_str,      # 规格图片（URL文本）
            main_images_url_str,    # 图片（URL文本）
            weight,                 # 重量
            stock_str,              # 库存（||分隔）
            site                    # 站点
        ]
        
        return row

    def generate_excel(self, results):
        dest = self.out_path.get()
        wb = Workbook()
        ws = wb.active
        headers = ['产品ID', '标题', '商品描述', '分类ID', '价格', '规格一', '规格一内容', 
                   '规格二', '规格二内容', '规格图片', '图片', '重量', '库存', '站点']
        ws.append(headers)

        # 美化表头
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid")
            cell.font = Font(bold=True, color="1890FF")

        # 写入数据行
        for r_idx, row_data in enumerate(results, start=2):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 设置行高（稍微低一点，因为不需要嵌入图片）
            ws.row_dimensions[r_idx].height = 50

        # 调整列宽
        ws.column_dimensions['A'].width = 12  # 产品ID
        ws.column_dimensions['B'].width = 40  # 标题
        ws.column_dimensions['C'].width = 30  # 商品描述
        ws.column_dimensions['D'].width = 10  # 分类ID
        ws.column_dimensions['E'].width = 20  # 价格
        ws.column_dimensions['F'].width = 12  # 规格一
        ws.column_dimensions['G'].width = 30  # 规格一内容
        ws.column_dimensions['H'].width = 12  # 规格二
        ws.column_dimensions['I'].width = 30  # 规格二内容
        ws.column_dimensions['J'].width = 60  # 规格图片URL
        ws.column_dimensions['K'].width = 60  # 图片URL
        ws.column_dimensions['L'].width = 10  # 重量
        ws.column_dimensions['M'].width = 20  # 库存
        ws.column_dimensions['N'].width = 8   # 站点
        
        try:
            self.progress_text.set("正在保存Excel文件...")
            wb.save(dest)
            self.logger.log(f"任务圆满完成！Excel: {dest}", "SUCCESS")
            self.progress_text.set("✓ 全部完成")
            
            # 生成统计报告
            report = f"数据已成功导出至 Excel\n\n"
            report += f"商品数量: {len(results)}\n"
            report += f"导出格式: 商品上架格式（网络链接版）\n"
            report += f"图片格式: URL文本（用||分隔）\n"
            report += f"\n说明：图片以URL形式保存，不下载到本地"
            
            # 使用 after 在主线程中显示对话框
            self.root.after(0, lambda: messagebox.showinfo("完成", report))
            
        except PermissionError:
            self.logger.log("错误：文件被占用，无法写入！", "ERROR")
            self.progress_text.set("✗ 保存失败")
            # 使用 after 在主线程中显示错误对话框
            self.root.after(0, lambda: messagebox.showerror("错误", "文件可能已在 Excel 中打开，请先关闭！"))


if __name__ == "__main__":
    root = tk.Tk()
    app = PddParserV2(root)
    root.mainloop()
