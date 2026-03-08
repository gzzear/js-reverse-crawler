import glob
import json
import os
import re
import threading
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk, scrolledtext
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image

# --- 工业级视觉规范 ---
COLOR_NAVY_DARK = "#001529"  # 导航深蓝
COLOR_ANT_BLUE = "#1890ff"  # 品牌蓝
COLOR_BG_LIGHT = "#f0f2f5"  # 页面背景
COLOR_WHITE = "#ffffff"  # 卡片背景

# 极速连接池
http_session = requests.Session()
adapter = requests.adapters.HTTPAdapter(pool_connections=100, pool_maxsize=100)
http_session.mount('https://', adapter)
http_session.mount('http://', adapter)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"}


class ProfessionalLogger:
    def __init__(self, text_widget):
        self.widget = text_widget
        self.widget.tag_config("TIME", foreground="#8c8c8c")
        self.widget.tag_config("INFO", foreground="#1890ff")
        self.widget.tag_config("SUCCESS", foreground="#52c41a", font=("Segoe UI", 9, "bold"))
        self.widget.tag_config("WARNING", foreground="#faad14")
        self.widget.tag_config("ERROR", foreground="#ff4d4f", font=("Segoe UI", 9, "bold"))

    def log(self, message, level="INFO"):
        now = datetime.now().strftime("%H:%M:%S")
        self.widget.config(state=tk.NORMAL)
        self.widget.insert(tk.END, f"[{now}] ", "TIME")
        self.widget.insert(tk.END, f"[{level}] ", level)
        self.widget.insert(tk.END, f"{message}\n")
        self.widget.see(tk.END)
        self.widget.config(state=tk.DISABLED)

    def clear(self):
        self.widget.config(state=tk.NORMAL)
        self.widget.delete(1.0, tk.END)
        self.widget.config(state=tk.DISABLED)


def clean_json_content(content):
    """鲁棒性清洗 JSON 杂质"""
    content = content.strip()
    match = re.search(r'(\{.*\})', content, re.DOTALL)
    return match.group(1) if match else content


class PddParserV2:
    def __init__(self, root):
        self.root = root
        self.root.title("PDD 数据解析专家 v4.0 (多格式导出版)")
        self.root.geometry("1000x750")
        self.root.configure(bg=COLOR_BG_LIGHT)

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass

        self.setup_styles()
        self.build_ui()
        self.logger = ProfessionalLogger(self.log_area)

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Card.TFrame", background=COLOR_WHITE, relief="flat")
        style.configure("TLabel", background=COLOR_WHITE, font=("Microsoft YaHei", 10))
        style.configure("Header.TLabel", font=("Microsoft YaHei", 18, "bold"), background=COLOR_NAVY_DARK,
                        foreground="white")
        style.configure("Primary.TButton", font=("Microsoft YaHei", 10, "bold"), padding=8)
        style.map("Primary.TButton", background=[('pressed', '#096dd9'), ('active', COLOR_ANT_BLUE)],
                  foreground=[('!disabled', 'white')])

    def build_ui(self):
        # 1. 顶部标题栏
        nav = tk.Frame(self.root, bg=COLOR_NAVY_DARK, height=65)
        nav.pack(fill=tk.X)
        nav.pack_propagate(False)
        ttk.Label(nav, text=" PDD DATA PARSER V4 ", style="Header.TLabel").pack(side=tk.LEFT, padx=30)
        ttk.Label(nav, text=" 多格式导出 | 支持自定义模板 ", font=("Microsoft YaHei", 9), background=COLOR_NAVY_DARK,
                  foreground="#595959").pack(side=tk.LEFT, pady=(8, 0))

        # 2. 主卡片容器
        container = ttk.Frame(self.root, padding=30)
        container.pack(fill=tk.BOTH, expand=True)

        cfg_card = ttk.Frame(container, style="Card.TFrame", padding=25)
        cfg_card.pack(fill=tk.X, pady=(0, 20))

        # 输入路径配置
        f1 = ttk.Frame(cfg_card, style="Card.TFrame")
        f1.pack(fill=tk.X)
        ttk.Label(f1, text="数据源路径", font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.in_path = tk.StringVar()
        ttk.Entry(f1, textvariable=self.in_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=20)

        btn_box = ttk.Frame(f1, style="Card.TFrame")
        btn_box.pack(side=tk.LEFT)
        ttk.Button(btn_box, text="单文件", width=8, command=self.select_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_box, text="文件夹", width=8, command=self.select_dir).pack(side=tk.LEFT)

        # 导出格式选择
        f2 = ttk.Frame(cfg_card, style="Card.TFrame")
        f2.pack(fill=tk.X, pady=(15, 0))
        ttk.Label(f2, text="导出格式    ", font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.export_format = tk.StringVar(value="standard")
        ttk.Radiobutton(f2, text="标准格式", variable=self.export_format, value="standard").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(f2, text="导入模板格式", variable=self.export_format, value="template").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(f2, text="拼多多格式", variable=self.export_format, value="pinduoduo", command=self.on_format_change).pack(side=tk.LEFT, padx=10)
        
        # 图片保存目录（仅拼多多格式需要）
        self.img_dir_frame = ttk.Frame(cfg_card, style="Card.TFrame")
        self.img_dir = tk.StringVar()
        ttk.Label(self.img_dir_frame, text="图片目录    ", font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        ttk.Entry(self.img_dir_frame, textvariable=self.img_dir).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=20)
        ttk.Button(self.img_dir_frame, text="浏览...", width=10, command=self.select_img_dir).pack(side=tk.LEFT)
        
        # 导出路径配置
        f3 = ttk.Frame(cfg_card, style="Card.TFrame")
        f3.pack(fill=tk.X, pady=(15, 0))
        ttk.Label(f3, text="导出 Excel  ", font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.out_path = tk.StringVar()
        ttk.Entry(f3, textvariable=self.out_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=20)
        ttk.Button(f3, text="浏览...", width=10, command=self.select_out).pack(side=tk.LEFT)

        # 3. 控制面板
        ctrl_fm = ttk.Frame(container)
        ctrl_fm.pack(fill=tk.X, pady=(10, 20))
        self.run_btn = ttk.Button(ctrl_fm, text="🚀 开始解析并导出", style="Primary.TButton",
                                  command=self.start_task)
        self.run_btn.pack(side=tk.LEFT)
        ttk.Button(ctrl_fm, text="🗑 清空日志", command=lambda: self.logger.clear()).pack(side=tk.RIGHT)

        # 4. 进度条
        progress_frame = ttk.Frame(container)
        progress_frame.pack(fill=tk.X, pady=(10, 5))

        ttk.Label(progress_frame, text="解析进度:", font=("Microsoft YaHei", 9)).pack(anchor=tk.W)
        self.pbar = ttk.Progressbar(progress_frame, mode='determinate')
        self.pbar.pack(fill=tk.X, pady=(2, 10))
        self.progress_text = tk.StringVar(value="就绪")
        ttk.Label(progress_frame, textvariable=self.progress_text, font=("Microsoft YaHei", 8),
                  foreground="#8c8c8c").pack(anchor=tk.W)
        
        # 图片下载进度条
        self.img_pbar = ttk.Progressbar(progress_frame, mode='determinate')
        self.img_progress_text = tk.StringVar(value="")

        # 5. 日志区
        self.log_area = scrolledtext.ScrolledText(container, height=18, font=("Consolas", 10), bg="#ffffff",
                                                  borderwidth=0, highlightthickness=1, highlightbackground="#d9d9d9",
                                                  state=tk.DISABLED)
        self.log_area.pack(fill=tk.BOTH, expand=True, pady=10)

    def select_file(self):
        p = filedialog.askopenfilename(filetypes=[("数据包", "*.txt *.json"), ("所有文件", "*.*")])
        if p: self.in_path.set(p)

    def select_dir(self):
        p = filedialog.askdirectory()
        if p: self.in_path.set(p)

    def select_out(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p: self.out_path.set(p)
    
    def select_img_dir(self):
        p = filedialog.askdirectory(title="选择图片保存目录")
        if p: self.img_dir.set(p)
    
    def on_format_change(self):
        """格式切换时显示/隐藏图片目录选择"""
        if self.export_format.get() == "pinduoduo":
            self.img_dir_frame.pack(fill=tk.X, pady=(15, 0))
        else:
            self.img_dir_frame.pack_forget()

    def start_task(self):
        if not self.in_path.get() or not self.out_path.get():
            messagebox.showwarning("提示", "请先配置输入源和保存位置")
            return
        
        # 如果是拼多多格式，需要检查图片目录
        if self.export_format.get() == "pinduoduo" and not self.img_dir.get():
            messagebox.showwarning("提示", "拼多多格式需要指定图片保存目录")
            return
            
        self.run_btn.config(state=tk.DISABLED)
        threading.Thread(target=self.work_thread, daemon=True).start()

    def work_thread(self):
        self.logger.log(">>> 任务启动，开始扫描...", "INFO")
        inp = self.in_path.get()

        # 兼容单文件与文件夹
        if os.path.isfile(inp):
            target_files = [inp]
        else:
            files = glob.glob(os.path.join(inp, "**", "*.*"), recursive=True)
            target_files = [f for f in files if f.lower().endswith(('.json', '.txt'))]

        if not target_files:
            self.logger.log("错误：未发现有效文件包", "ERROR")
            self.run_btn.config(state=tk.NORMAL)
            return

        self.pbar['maximum'] = len(target_files)
        self.pbar['value'] = 0
        data_rows = []

        for idx, f_path in enumerate(target_files):
            # 更新文件解析进度
            self.progress_text.set(f"正在解析: {os.path.basename(f_path)} ({idx + 1}/{len(target_files)})")

            # 智能解析
            res = self.smart_parse(f_path)
            if res:
                data_rows.append(res)
                self.logger.log(f"解析成功: {os.path.basename(f_path)}", "SUCCESS")
            else:
                self.logger.log(f"跳过非详情包: {os.path.basename(f_path)}", "WARNING")

            self.pbar['value'] = idx + 1
            self.root.update_idletasks()

        if data_rows:
            self.logger.log(f"解析完毕。有效数据: {len(data_rows)} / 总文件: {len(target_files)}", "INFO")
            self.progress_text.set(f"解析完成，开始生成Excel...")
            self.generate_excel(data_rows)
        else:
            self.logger.log("解析结果为空，任务提前结束", "ERROR")
            self.progress_text.set("解析失败")
        self.run_btn.config(state=tk.NORMAL)

    def smart_parse(self, path):
        """核心：智能适配多种 PDD 接口结构 - V4多格式版"""
        content = None
        for enc in ['utf-8-sig', 'utf-8', 'gbk', 'utf-16', 'latin1']:
            try:
                with open(path, 'r', encoding=enc) as f:
                    raw_content = f.read()
                    content = json.loads(clean_json_content(raw_content))
                    break
            except:
                continue
        if not content: return None

        # 智能层级探测（兼容两种格式）
        # 格式1: store.initDataObj.goods
        # 格式2: 直接 goods
        store_data = content.get('store', {})
        init_data = store_data.get('initDataObj', {})
        goods = init_data.get('goods') or content.get('goods')
        if not goods: return None

        def sget(d, *keys):
            for k in keys:
                if k in d: return d[k]
            return None

        # 基础字段
        goods_name = sget(goods, 'goodsName', 'goods_name', 'shareDesc', 'share_desc') or 'NA'
        goods_id = sget(goods, 'goodsID', 'goods_id') or '0'
        cat_id = str(sget(goods, 'catID', 'cat_id') or '')

        # 商品描述：从属性中构建
        props = goods.get('goodsProperty', []) or goods.get('goods_property', [])
        desc_lines = []
        for p in props:
            if isinstance(p, dict) and p.get('key'):
                key = p.get('key')
                values = p.get('values', [])
                if values:
                    value = values[0] if isinstance(values, list) else values
                    desc_lines.append(f"{key}：{value}")
        goods_desc = '\n'.join(desc_lines)

        # SKU处理（兼容两种字段名）
        skus = goods.get('skus', []) or content.get('sku', [])

        # 提取规格维度
        spec_keys = {}  # {spec_key_id: spec_key_name}
        spec_values = {}  # {spec_key_id: [spec_values]}

        for sku in skus:
            if not isinstance(sku, dict):
                continue
            specs = sku.get('specs', [])
            for spec in specs:
                if isinstance(spec, dict):
                    spec_key_id = spec.get('spec_key_id') or spec.get('specKeyId')
                    spec_key = spec.get('spec_key') or spec.get('specKey', '')
                    spec_value = spec.get('spec_value') or spec.get('specValue', '')

                    if spec_key_id:
                        spec_keys[spec_key_id] = spec_key
                        if spec_key_id not in spec_values:
                            spec_values[spec_key_id] = []
                        if spec_value and spec_value not in spec_values[spec_key_id]:
                            spec_values[spec_key_id].append(spec_value)

        # 获取规格一和规格二
        spec_key_ids = sorted(spec_keys.keys())
        spec1_name = spec_keys.get(spec_key_ids[0], '') if len(spec_key_ids) > 0 else ''
        spec1_values = '||'.join(spec_values.get(spec_key_ids[0], [])) if len(spec_key_ids) > 0 else ''
        spec2_name = spec_keys.get(spec_key_ids[1], '') if len(spec_key_ids) > 1 else ''
        spec2_values = '||'.join(spec_values.get(spec_key_ids[1], [])) if len(spec_key_ids) > 1 else ''

        # 价格、库存、规格图片（按SKU顺序）
        prices = []
        stocks = []
        sku_images = []

        for sku in skus:
            if not isinstance(sku, dict):
                continue

            # 价格处理（兼容多种格式）
            group_price = sku.get('groupPrice') or sku.get('group_price')
            sku_price = sku.get('skuPrice') or sku.get('sku_price', 0)

            # 判断价格格式
            if isinstance(group_price, str) and group_price:
                # 已经是字符串格式的元（如 "69.9"）
                price = group_price
            elif isinstance(group_price, (int, float)) and group_price > 0:
                # 数字格式，判断是否需要除以100
                if group_price > 1000:  # 大于1000认为是分
                    price = str(float(group_price) / 100)
                else:
                    price = str(float(group_price))
            elif sku_price:
                # 使用sku_price，通常是分
                price = str(float(sku_price) / 100) if sku_price else '0'
            else:
                price = '0'

            prices.append(price)

            # 库存
            quantity = sku.get('quantity', 0)
            stocks.append(str(quantity))

            # SKU图片URL
            thumb_url = sku.get('thumbUrl') or sku.get('thumb_url', '')
            sku_images.append(thumb_url)

        price_str = '||'.join(prices)
        stock_str = '||'.join(stocks)

        # 主图URL提取 - 按优先级尝试多个字段（只提取主图，不包含详情图）
        m_urls_original = []
        use_carousel = False  # 标记是否使用了轮播图
        
        # 先收集所有可能的主图来源
        # 优先级1：ui.carousel_section.picture_list（轮播图，最准确的主图来源）
        ui_data = content.get('ui', {})
        carousel_section = ui_data.get('carousel_section', {})
        picture_list = carousel_section.get('picture_list', [])
        if picture_list:
            m_urls_original = [pic.get('url') for pic in picture_list if isinstance(pic, dict) and pic.get('url')]
            use_carousel = True  # 使用了轮播图
        
        # 优先级2：viewImageData（URL字符串列表）
        if not m_urls_original:
            view_image_data = goods.get('viewImageData', [])
            if view_image_data and isinstance(view_image_data, list):
                m_urls_original = view_image_data
        
        # 优先级3：gallery
        if not m_urls_original:
            gallery = goods.get('gallery', [])
            if gallery:
                m_urls_original = [i.get('url') if isinstance(i, dict) else i for i in gallery]

        # 优先级4：topGallery（字典列表，需要提取url字段）
        if not m_urls_original:
            top_gallery = goods.get('topGallery', [])
            if top_gallery:
                m_urls_original = [i.get('url') if isinstance(i, dict) else i for i in top_gallery]
        
        # 优先级5：thumbUrl作为兜底
        if not m_urls_original:
            thumb_url = sget(goods, 'thumbUrl', 'thumb_url', 'hd_thumb_url')
            if thumb_url:
                m_urls_original = [thumb_url]
        
        # 如果使用了轮播图，直接使用不过滤（轮播图就是主图）
        if use_carousel:
            m_urls = m_urls_original
        else:
            # 其他来源需要过滤掉详情图
            decoration = goods.get('decoration', [])
            decoration_urls = set()
            for dec in decoration:
                if isinstance(dec, dict) and dec.get('type') == 'image':
                    for content_item in dec.get('contents', []):
                        if isinstance(content_item, dict):
                            img_url = content_item.get('img_url') or content_item.get('imgUrl')
                            if img_url:
                                decoration_urls.add(img_url)
            
            # 过滤：保留主图中不在详情图中的URL
            if decoration_urls:
                m_urls_filtered = [url for url in m_urls_original if url not in decoration_urls]
                # 如果过滤后为空，使用thumbUrl作为兜底（不使用详情图）
                if m_urls_filtered:
                    m_urls = m_urls_filtered
                else:
                    # 所有主图都是详情图，使用thumbUrl兜底
                    thumb_url = sget(goods, 'thumbUrl', 'thumb_url', 'hd_thumb_url')
                    m_urls = [thumb_url] if thumb_url else []
            else:
                # 没有详情图，直接使用原始主图列表
                m_urls = m_urls_original

        # 重量提取（优先级：goods.weight > skus[].weight > goodsProperty）
        weight = '0'
        goods_weight = goods.get('weight', 0)
        if goods_weight and goods_weight != 0:
            weight = str(goods_weight)
        else:
            # 尝试从SKU中获取（取第一个非0的）
            for sku in skus:
                sku_weight = sku.get('weight', 0)
                if sku_weight and sku_weight != 0:
                    weight = str(sku_weight)
                    break

            # 如果还是0，尝试从goodsProperty中提取
            if weight == '0':
                for prop in props:
                    if '重量' in prop.get('key', ''):
                        values = prop.get('values', [])
                        if values and values[0] not in ['其他', '其他）', '']:
                            weight = str(values[0])
                            break

        # 站点/国家提取（优先级：goods.country > goodsProperty产地 > 默认CN）
        site = 'CN'  # 默认中国

        # 方法1：从 goods.country 获取
        country = goods.get('country', '')
        if country and country.strip():
            site = country.strip()
        else:
            # 方法2：从 goodsProperty 中提取产地或发货地
            for prop in props:
                key = prop.get('key', '')
                if '产地' in key or '发货地' in key:
                    values = prop.get('values', [])
                    if values and values[0]:
                        origin = values[0]
                        # 简单映射：中国大陆->CN, 其他国家可以扩展
                        if '中国' in origin or '大陆' in origin:
                            site = 'CN'
                        elif '泰国' in origin:
                            site = 'TH'
                        elif '美国' in origin:
                            site = 'US'
                        elif '日本' in origin:
                            site = 'JP'
                        elif '韩国' in origin:
                            site = 'KR'
                        else:
                            # 如果是省份（如"河北省"），保持CN
                            if '省' in origin or '市' in origin:
                                site = 'CN'
                            else:
                                site = origin[:2].upper()  # 取前两个字符作为代码
                        break

        # 提取详情图（decoration中的图片）
        decoration = goods.get('decoration', [])
        detail_images = []
        for dec in decoration:
            if isinstance(dec, dict) and dec.get('type') == 'image':
                for content_item in dec.get('contents', []):
                    if isinstance(content_item, dict):
                        img_url = content_item.get('img_url') or content_item.get('imgUrl')
                        if img_url:
                            detail_images.append(img_url)
        
        # 如果decoration为空，尝试从detailGallery获取详情图
        if not detail_images:
            detail_gallery = goods.get('detailGallery', [])
            if detail_gallery:
                for item in detail_gallery:
                    if isinstance(item, dict):
                        url = item.get('url') or item.get('image_url') or item.get('img_url')
                        if url:
                            detail_images.append(url)
                    elif isinstance(item, str):
                        detail_images.append(item)
        
        # 整理图片URL字符串
        sku_image_url_str = '||'.join([img for img in sku_images if img])
        main_images_url_str = '||'.join([url for url in m_urls if url])
        detail_images_url_str = '||'.join([url for url in detail_images if url])

        # 构建标准格式行数据
        row = [
            str(goods_id),  # 产品ID
            goods_name,  # 标题
            goods_desc,  # 商品描述
            cat_id,  # 分类ID
            price_str,  # 价格（||分隔）
            spec1_name,  # 规格一
            spec1_values,  # 规格一内容
            spec2_name,  # 规格二
            spec2_values,  # 规格二内容
            sku_image_url_str,  # 规格图片（URL文本）
            main_images_url_str,  # 图片（URL文本）
            weight,  # 重量
            stock_str,  # 库存（||分隔）
            detail_images_url_str,  # 详情图片（URL文本）
            site  # 站点
        ]

        # 构建完整数据（用于模板格式和拼多多格式）
        full_data = {
            'goods_id': str(goods_id),
            'goods_name': goods_name,
            'goods_desc': goods_desc,
            'cat_id': cat_id,
            'main_images': [url for url in m_urls if url],  # 过滤空值
            'weight': weight,
            'site': site,
            'props': props,  # 商品属性
            'brand_id': str(sget(goods, 'brandId', 'brand_id') or ''),  # 品牌ID
            'sales_tip': sget(goods, 'sideSalesTip', 'side_sales_tip') or '',  # 已售出提示
            'review_num': 0,  # 评价数（需要从content.store.initDataObj.review.reviewNum获取）
            'skus': []  # SKU详情列表
        }
        
        # 尝试获取评价数
        try:
            review_data = content.get('store', {}).get('initDataObj', {}).get('review', {})
            full_data['review_num'] = review_data.get('reviewNum', 0)
        except:
            pass
        
        # 整理每个SKU的详细信息
        for idx, sku in enumerate(skus):
            if not isinstance(sku, dict):
                continue
            
            # 提取该SKU的规格值
            sku_specs = []
            specs = sku.get('specs', [])
            for spec in specs:
                if isinstance(spec, dict):
                    spec_key = spec.get('spec_key') or spec.get('specKey', '')
                    spec_value = spec.get('spec_value') or spec.get('specValue', '')
                    if spec_key and spec_value:
                        sku_specs.append(f"{spec_key}:{spec_value}")
            
            # 价格
            group_price = sku.get('groupPrice') or sku.get('group_price')
            sku_price = sku.get('skuPrice') or sku.get('sku_price', 0)
            
            if isinstance(group_price, str) and group_price:
                price = group_price
            elif isinstance(group_price, (int, float)) and group_price > 0:
                if group_price > 1000:
                    price = str(float(group_price) / 100)
                else:
                    price = str(float(group_price))
            elif sku_price:
                price = str(float(sku_price) / 100) if sku_price else '0'
            else:
                price = '0'
            
            # SKU库存
            quantity = sku.get('quantity', 0)
            
            # SKU图片
            thumb_url = sku.get('thumbUrl') or sku.get('thumb_url', '')
            
            # SKU重量
            sku_weight = sku.get('weight', 0)
            if sku_weight and sku_weight != 0:
                sku_weight_val = str(sku_weight)
            else:
                sku_weight_val = weight
            
            # SKU编号
            sku_id = sku.get('skuId') or sku.get('sku_id', '')
            
            full_data['skus'].append({
                'sku_id': str(sku_id) if sku_id else '',
                'specs': sku_specs,  # ["颜色:红色", "尺码:XL"]
                'price': price,
                'stock': str(quantity),
                'image': thumb_url,
                'weight': sku_weight_val
            })

        return (row, full_data)

    def generate_excel(self, results):
        """根据选择的格式生成Excel"""
        export_format = self.export_format.get()
        
        if export_format == "template":
            self.generate_template_format(results)
        elif export_format == "pinduoduo":
            self.generate_pinduoduo_format(results)
        else:
            self.generate_standard_format(results)
    
    def generate_standard_format(self, results):
        """生成标准格式Excel"""
        dest = self.out_path.get()
        wb = Workbook()
        ws = wb.active
        headers = ['产品ID', '标题', '商品描述', '分类ID', '价格', '规格一', '规格一内容',
                   '规格二', '规格二内容', '规格图片', '图片', '重量', '库存', '详情图片', '站点']
        ws.append(headers)

        # 美化表头
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid")
            cell.font = Font(bold=True, color="1890FF")

        # 写入数据行
        for r_idx, (row_data, full_data) in enumerate(results, start=2):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            # 设置行高
            ws.row_dimensions[r_idx].height = 50

        # 调整列宽
        ws.column_dimensions['A'].width = 12  # 产品ID
        ws.column_dimensions['B'].width = 40  # 标题
        ws.column_dimensions['C'].width = 30  # 商品描述
        ws.column_dimensions['D'].width = 10  # 分类ID
        ws.column_dimensions['E'].width = 20  # 价格
        ws.column_dimensions['F'].width = 12  # 规格一
        ws.column_dimensions['G'].width = 30  # 规格一内容
        ws.column_dimensions['H'].width = 12  # 规格二
        ws.column_dimensions['I'].width = 30  # 规格二内容
        ws.column_dimensions['J'].width = 60  # 规格图片
        ws.column_dimensions['K'].width = 60  # 图片
        ws.column_dimensions['L'].width = 10  # 重量
        ws.column_dimensions['M'].width = 20  # 库存
        ws.column_dimensions['N'].width = 60  # 详情图片
        ws.column_dimensions['O'].width = 8   # 站点

        try:
            self.progress_text.set("正在保存Excel文件...")
            wb.save(dest)
            self.logger.log(f"任务圆满完成！Excel: {dest}", "SUCCESS")
            self.progress_text.set("✓ 全部完成")

            report = f"数据已成功导出至 Excel\n\n"
            report += f"商品数量: {len(results)}\n"
            report += f"导出格式: 标准格式\n"
            report += f"图片格式: URL文本（用||分隔）"

            self.root.after(0, lambda: messagebox.showinfo("完成", report))

        except PermissionError:
            self.logger.log("错误：文件被占用，无法写入！", "ERROR")
            self.progress_text.set("✗ 保存失败")
            self.root.after(0, lambda: messagebox.showerror("错误", "文件可能已在 Excel 中打开，请先关闭！"))
    
    def generate_template_format(self, results):
        """生成导入模板格式Excel（一个商品一行）"""
        dest = self.out_path.get()
        wb = Workbook()
        ws = wb.active
        
        # 模板表头（20列）
        headers = [
            '产品主编号', '产品名称', '货币类型', '产品主图', '货源链接', 
            '货源平台', '货源ID', '详情描述', '详情图', '货源类目',
            '自定义属性', '产品视频', 'SKU规格1', 'SKU规格2', '平台SKU',
            'SKU售价', 'SKU图片', 'SKU库存', 'SKU重量(KG)', 'SKU尺寸(CM)'
        ]
        ws.append(headers)

        # 美化表头
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据（一个商品一行，多SKU用分隔符合并）
        total_sku_count = 0
        for row_data, full_data in results:
            goods_id = full_data['goods_id']
            goods_name = full_data['goods_name']
            goods_desc = full_data['goods_desc']
            main_images = full_data['main_images']
            skus = full_data['skus']
            props = full_data['props']
            
            # 构建产品主图（用中文逗号分隔）
            main_images_str = '，'.join(main_images[:5]) if main_images else ''
            
            # 构建详情图（用中文逗号分隔）
            detail_images_str = '，'.join(main_images) if main_images else ''
            
            # 构建自定义属性（用中文分号分隔）
            custom_attrs = []
            for prop in props:
                if isinstance(prop, dict) and prop.get('key'):
                    key = prop.get('key')
                    values = prop.get('values', [])
                    if values:
                        value = values[0] if isinstance(values, list) else values
                        custom_attrs.append(f"{key}:{value}")
            custom_attrs_str = '；'.join(custom_attrs)
            
            # 如果没有SKU，至少输出一行
            if not skus:
                skus = [{'sku_id': '', 'specs': [], 'price': '0', 'stock': '0', 'image': '', 'weight': full_data['weight']}]
            
            total_sku_count += len(skus)
            
            # 合并所有SKU的信息到一行（用中文逗号分隔）
            sku_spec1_list = []
            sku_spec2_list = []
            sku_id_list = []
            sku_price_list = []
            sku_image_list = []
            sku_stock_list = []
            sku_weight_list = []
            
            for sku in skus:
                # 提取规格1和规格2
                sku_spec1 = sku['specs'][0] if len(sku['specs']) > 0 else ''
                sku_spec2 = sku['specs'][1] if len(sku['specs']) > 1 else ''
                
                sku_spec1_list.append(sku_spec1)
                sku_spec2_list.append(sku_spec2)
                sku_id_list.append(sku['sku_id'])
                sku_price_list.append(sku['price'])
                sku_image_list.append(sku['image'] if sku['image'] else '')
                sku_stock_list.append(sku['stock'])
                sku_weight_list.append(sku['weight'])
            
            # 货源链接（拼多多商品链接）
            source_link = f"https://mobile.yangkeduo.com/goods.html?goods_id={goods_id}"
            
            # 一个商品一行
            row = [
                goods_id,                              # A: 产品主编号
                goods_name,                            # B: 产品名称
                'CNY',                                 # C: 货币类型
                main_images_str,                       # D: 产品主图
                source_link,                           # E: 货源链接
                '拼多多',                               # F: 货源平台
                goods_id,                              # G: 货源ID
                goods_desc,                            # H: 详情描述
                detail_images_str,                     # I: 详情图
                '',                                    # J: 货源类目（暂时为空）
                custom_attrs_str,                      # K: 自定义属性
                '',                                    # L: 产品视频（暂时为空）
                '，'.join(sku_spec1_list),             # M: SKU规格1（多个用逗号分隔）
                '，'.join(sku_spec2_list),             # N: SKU规格2（多个用逗号分隔）
                '，'.join(sku_id_list),                # O: 平台SKU（多个用逗号分隔）
                '，'.join(sku_price_list),             # P: SKU售价（多个用逗号分隔）
                '，'.join(sku_image_list),             # Q: SKU图片（多个用逗号分隔）
                '，'.join(sku_stock_list),             # R: SKU库存（多个用逗号分隔）
                '，'.join(sku_weight_list),            # S: SKU重量(KG)（多个用逗号分隔）
                ''                                     # T: SKU尺寸(CM)（暂时为空）
            ]
            
            ws.append(row)
            total_sku_count += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15  # 产品主编号
        ws.column_dimensions['B'].width = 30  # 产品名称
        ws.column_dimensions['C'].width = 10  # 货币类型
        ws.column_dimensions['D'].width = 50  # 产品主图
        ws.column_dimensions['E'].width = 50  # 货源链接
        ws.column_dimensions['F'].width = 12  # 货源平台
        ws.column_dimensions['G'].width = 15  # 货源ID
        ws.column_dimensions['H'].width = 40  # 详情描述
        ws.column_dimensions['I'].width = 50  # 详情图
        ws.column_dimensions['J'].width = 15  # 货源类目
        ws.column_dimensions['K'].width = 40  # 自定义属性
        ws.column_dimensions['L'].width = 40  # 产品视频
        ws.column_dimensions['M'].width = 20  # SKU规格1
        ws.column_dimensions['N'].width = 20  # SKU规格2
        ws.column_dimensions['O'].width = 15  # 平台SKU
        ws.column_dimensions['P'].width = 12  # SKU售价
        ws.column_dimensions['Q'].width = 40  # SKU图片
        ws.column_dimensions['R'].width = 10  # SKU库存
        ws.column_dimensions['S'].width = 12  # SKU重量
        ws.column_dimensions['T'].width = 15  # SKU尺寸
        
        # 设置所有数据行的对齐方式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        try:
            self.progress_text.set("正在保存Excel文件...")
            wb.save(dest)
            self.logger.log(f"任务圆满完成！Excel: {dest}", "SUCCESS")
            self.progress_text.set("✓ 全部完成")

            report = f"数据已成功导出至 Excel\n\n"
            report += f"商品数量: {len(results)}\n"
            report += f"SKU数量: {total_sku_count}\n"
            report += f"导出格式: 导入模板格式（按SKU展开）\n"
            report += f"图片分隔符: 中文逗号（，）\n"
            report += f"属性分隔符: 中文分号（；）"

            self.root.after(0, lambda: messagebox.showinfo("完成", report))

        except PermissionError:
            self.logger.log("错误：文件被占用，无法写入！", "ERROR")
            self.progress_text.set("✗ 保存失败")
            self.root.after(0, lambda: messagebox.showerror("错误", "文件可能已在 Excel 中打开，请先关闭！"))
    
    def generate_pinduoduo_format(self, results):
        """生成拼多多格式Excel（35列，一个商品一行，图片下载到本地）"""
        dest = self.out_path.get()
        img_base_dir = self.img_dir.get()
        wb = Workbook()
        ws = wb.active
        
        # 35列表头
        headers = [
            'ID', '名稱', '價格', '描述', '圖片', '數量', '重量', '分組', '網站', '特品',
            '品牌', '類別編號', '商品選項', '出貨天數', '保存狀況', '商品主貨號', '類別屬性', 
            '多件優惠', '包裹尺寸', '運費', '帳號', '尺碼圖表', '最低購買數量', '紀錄', '網址',
            '瀏覽量', '已售出', '關注', '評價', '商品編號', '創建時間', '修改時間', 
            '上架方式', '危險物品', '選中'
        ]
        ws.append(headers)
        
        # 美化表头
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 收集所有需要下载的图片（主图 + SKU图）
        self.logger.log("开始收集图片URL...", "INFO")
        download_tasks = []  # (url, goods_id, img_type, img_index)
        
        for row_data, full_data in results:
            goods_id = full_data['goods_id']
            main_images = full_data['main_images']
            skus = full_data['skus']
            
            # 添加调试信息
            if not main_images:
                self.logger.log(f"商品 {goods_id} 没有主图URL", "WARNING")
            
            # 收集主图
            for idx, url in enumerate(main_images, 1):
                if url:
                    download_tasks.append((url, goods_id, 'main', idx))
            
            # 收集SKU图
            for idx, sku in enumerate(skus, 1):
                sku_image = sku.get('image', '')
                if sku_image:
                    download_tasks.append((sku_image, goods_id, 'sku', idx))
        
        self.logger.log(f"共需下载 {len(download_tasks)} 张图片（来自{len(results)}个商品，包含主图和SKU图）", "INFO")
        
        # 批量下载图片
        image_path_map = {}  # {(goods_id, url): local_path}
        if download_tasks:
            image_path_map = self.bulk_download_images(download_tasks, img_base_dir)
        
        # 写入数据
        self.progress_text.set("正在生成Excel数据...")
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        total_images_in_excel = 0  # 统计实际写入Excel的图片数
        
        for idx, (row_data, full_data) in enumerate(results, start=1):
            goods_id = full_data['goods_id']
            goods_name = full_data['goods_name']
            goods_desc = full_data['goods_desc']
            main_images = full_data['main_images']
            skus = full_data['skus']
            props = full_data['props']
            weight = full_data['weight']
            cat_id = full_data['cat_id']
            
            # 1. ID
            row_id = idx
            
            # 2. 名稱
            name = goods_name
            
            # 3. 價格 (取最低价)
            prices = [float(sku['price']) for sku in skus if sku['price'] != '0']
            price = str(min(prices)) if prices else '0'
            
            # 4. 描述 (HTML格式，从goodsProperty生成)
            desc_html = self.build_desc_html(goods_desc, props)
            
            # 5. 圖片 (本地路径，用|分隔，包含主图和SKU图)
            local_image_paths = []
            
            # 添加主图
            for url in main_images:
                if url:
                    local_path = image_path_map.get((goods_id, url), '')
                    if local_path:
                        local_image_paths.append(local_path)
            
            # 添加SKU图
            for sku in skus:
                sku_image = sku.get('image', '')
                if sku_image:
                    local_path = image_path_map.get((goods_id, sku_image), '')
                    if local_path:
                        local_image_paths.append(local_path)
            
            image_str = '|'.join(local_image_paths)
            
            # 统计
            total_images_in_excel += len(local_image_paths)
            
            # 如果没有图片路径，记录警告
            if not local_image_paths and (main_images or any(sku.get('image') for sku in skus)):
                self.logger.log(f"商品 {goods_id} 有图片URL但没有本地路径", "WARNING")
            
            # 6. 數量 (所有SKU库存总和)
            total_qty = sum([int(sku['stock']) for sku in skus])
            
            # 7. 重量
            weight_val = weight if weight != '0' else ''
            
            # 8. 分組
            group = '拼多多'
            
            # 9. 網站
            website = '拼多多'
            
            # 10. 特品
            special = False
            
            # 11. 品牌
            brand = full_data.get('brand_id', '')
            
            # 12. 類別編號
            category_no = cat_id
            
            # 13. 商品選項 (复杂JSON)
            sku_options_json = self.build_sku_options_json(skus)
            
            # 14-24. 默认值或空
            ship_days = ''
            storage = ''
            main_goods_no = goods_id
            category_attrs = ''
            multi_discount = self.extract_yellow_labels(skus)
            package_size = ''
            freight = ''
            account = ''
            size_chart = ''
            min_buy_qty = ''
            record = ''
            
            # 25. 網址
            url = f"https://mobile.yangkeduo.com/goods.html?goods_id={goods_id}"
            
            # 26-35. 其他字段
            views = 0
            sold = self.extract_sold_count(full_data.get('sales_tip', ''))
            follow = 0
            reviews = full_data.get('review_num', 0)
            goods_no = goods_id
            create_time = ''
            modify_time = ''
            shelf_method = ''
            dangerous = False
            selected = False
            
            row = [
                row_id, name, price, desc_html, image_str, total_qty, weight_val, group, 
                website, special, brand, category_no, sku_options_json, ship_days, storage,
                main_goods_no, category_attrs, multi_discount, package_size, freight, account,
                size_chart, min_buy_qty, record, url, views, sold, follow, reviews, goods_no,
                create_time, modify_time, shelf_method, dangerous, selected
            ]
            
            ws.append(row)
        
        # 调整列宽
        column_widths = [8, 35, 10, 40, 80, 10, 10, 12, 12, 8, 12, 12, 100, 10, 10, 15, 15, 
                        20, 15, 10, 12, 15, 12, 12, 50, 10, 10, 8, 10, 15, 18, 18, 12, 12, 8]
        for i, width in enumerate(column_widths, start=1):
            col_letter = chr(64 + i) if i <= 26 else f'A{chr(64 + i - 26)}'
            ws.column_dimensions[col_letter].width = width
        
        # 设置所有数据行的对齐方式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        try:
            self.progress_text.set("正在保存Excel文件...")
            wb.save(dest)
            self.logger.log(f"任务圆满完成！Excel: {dest}", "SUCCESS")
            self.progress_text.set("✓ 全部完成")

            report = f"数据已成功导出至 Excel\n\n"
            report += f"商品数量: {len(results)}\n"
            report += f"导出格式: 拼多多格式（35列）\n"
            report += f"图片下载: {len(image_path_map)}张（主图+SKU图）\n"
            report += f"Excel图片: {total_images_in_excel}个路径\n"
            report += f"图片目录: {img_base_dir}"
            
            self.root.after(0, lambda: messagebox.showinfo("完成", report))

        except PermissionError:
            self.logger.log("错误：文件被占用，无法写入！", "ERROR")
            self.progress_text.set("✗ 保存失败")
            self.root.after(0, lambda: messagebox.showerror("错误", "文件可能已在 Excel 中打开，请先关闭！"))
    
    def bulk_download_images(self, tasks, base_dir):
        """批量下载图片（主图 + SKU图）"""
        image_map = {}
        total = len(tasks)
        completed = 0
        
        # 显示图片下载进度条
        self.img_pbar.pack(fill=tk.X, pady=(10, 2))
        img_label = ttk.Label(self.img_pbar.master, textvariable=self.img_progress_text, 
                             font=("Microsoft YaHei", 8), foreground="#8c8c8c")
        img_label.pack(anchor=tk.W)
        
        self.img_pbar['maximum'] = total
        self.img_pbar['value'] = 0
        
        # 创建拼多多_pic目录
        pic_dir = os.path.join(base_dir, '拼多多_pic')
        os.makedirs(pic_dir, exist_ok=True)
        
        def download_one(task):
            url, goods_id, img_type, img_idx = task  # 增加img_type参数
            try:
                # 下载图片
                resp = http_session.get(url, headers=HEADERS, timeout=10)
                if resp.status_code == 200:
                    # 保存图片
                    img = Image.open(BytesIO(resp.content))
                    
                    # 转换为RGB（如果是RGBA）
                    if img.mode == 'RGBA':
                        img = img.convert('RGB')
                    
                    # 调整大小（保持比例，最大800x800）
                    img.thumbnail((800, 800), Image.Resampling.LANCZOS)
                    
                    # 生成文件名（区分主图和SKU图）
                    if img_type == 'main':
                        filename = f"{goods_id}-主图{img_idx:02d}.jpg"
                    else:  # sku
                        filename = f"{goods_id}-SKU{img_idx:02d}.jpg"
                    
                    local_path = os.path.join(pic_dir, filename)
                    
                    # 保存
                    img.save(local_path, 'JPEG', quality=85)
                    
                    # 返回相对路径（相对于Excel所在目录）
                    relative_path = f"\\拼多多_pic\\{filename}"
                    return (goods_id, url, relative_path)
            except Exception as e:
                self.logger.log(f"图片下载失败: {url[:50]}... - {str(e)}", "WARNING")
            return (goods_id, url, None)
        
        # 使用线程池并发下载
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = [executor.submit(download_one, task) for task in tasks]
            
            for future in as_completed(futures):
                goods_id, url, local_path = future.result()
                if local_path:
                    image_map[(goods_id, url)] = local_path
                
                completed += 1
                self.img_pbar['value'] = completed
                self.img_progress_text.set(f"图片下载进度: {completed}/{total}")
                self.root.update_idletasks()
        
        self.logger.log(f"图片下载完成: 成功 {len(image_map)}/{total}", "SUCCESS")
        
        # 隐藏图片进度条
        self.img_pbar.pack_forget()
        img_label.pack_forget()
        
        return image_map
    
    def build_desc_html(self, goods_desc, props):
        """构建HTML格式的商品描述"""
        html_parts = []
        
        # 从goodsProperty生成描述
        if props:
            for prop in props:
                if isinstance(prop, dict) and prop.get('key'):
                    key = prop.get('key', '')
                    values = prop.get('values', [])
                    if values:
                        value = values[0] if isinstance(values, list) else values
                        html_parts.append(f"<p><strong>{key}:</strong> {value}</p>")
        
        # 如果有goods_desc，添加到前面
        if goods_desc:
            html_parts.insert(0, f"<p>{goods_desc.replace(chr(10), '<br>')}</p>")
        
        return ''.join(html_parts) if html_parts else ''
    
    def build_sku_options_json(self, skus):
        """构建商品选项JSON（模仿模板格式）"""
        if not skus:
            return ''
        
        # 第一步：提取规格维度
        spec_keys = {}  # {spec_key: {spec_value: icon}}
        
        for sku in skus:
            for spec_text in sku['specs']:
                if ':' in spec_text:
                    spec_key, spec_value = spec_text.split(':', 1)
                    if spec_key not in spec_keys:
                        spec_keys[spec_key] = {}
                    if spec_value not in spec_keys[spec_key]:
                        spec_keys[spec_key][spec_value] = sku.get('image', '')
        
        # 构建dataRows
        data_rows = []
        for spec_key, spec_values_dict in spec_keys.items():
            spec_value_list = []
            for spec_value, icon in spec_values_dict.items():
                spec_value_list.append({
                    "name": spec_value,
                    "icon": icon
                })
            data_rows.append({
                "name": spec_key,
                "specValue": spec_value_list
            })
        
        # 构建SKU映射
        sku_mapping = {}
        for sku in skus:
            # 构建SKU组合键（如: "颜色/尺寸"）
            spec_parts = []
            for spec_text in sku['specs']:
                if ':' in spec_text:
                    spec_parts.append(spec_text.split(':', 1)[1])
            
            sku_key = '/'.join(spec_parts) if spec_parts else 'default'
            
            sku_mapping[sku_key] = {
                "price": sku['price'],
                "originalQty": sku['stock'],
                "barCode": ""
            }
        
        # 组合成最终JSON
        result = [
            {"dataRows": data_rows},
            sku_mapping
        ]
        
        return json.dumps(result, ensure_ascii=False)
    
    def extract_yellow_labels(self, skus):
        """提取优惠标签"""
        if not skus:
            return ''
        
        # 从第一个SKU提取优惠信息
        # 注意：full_data中没有存储yellowLabelList，这里返回空
        return ''
    
    def extract_sold_count(self, sales_tip):
        """从销售提示中提取数字"""
        if not sales_tip:
            return 0
        
        # 提取数字（如 "已拼1852件" -> 1852）
        match = re.search(r'(\d+)', sales_tip)
        return int(match.group(1)) if match else 0


if __name__ == "__main__":
    root = tk.Tk()
    app = PddParserV2(root)
    root.mainloop()
